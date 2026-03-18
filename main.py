# app.py
import sqlite3
from datetime import datetime, date
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import html
import streamlit.components.v1 as components
import calendar

from playwright.sync_api import sync_playwright



st.set_page_config(page_title="WLI Onboarding", layout="wide")

if "nav_page" not in st.session_state:
    st.session_state.nav_page = "📊 Today"

if "nav_request" not in st.session_state:
    st.session_state.nav_request = None

if "active_company_id" not in st.session_state:
    st.session_state.active_company_id = None

# ✅ Apply pending navigation BEFORE sidebar widgets are created
if st.session_state.nav_request:
    st.session_state.nav_page = st.session_state.nav_request
    st.session_state.nav_request = None

def set_active_company(company_id: int, go_page: str | None = None):
    st.session_state.active_company_id = int(company_id)
    if go_page:
        st.session_state.nav_request = go_page
        st.rerun()



st.markdown("""
<style>
/* ===== ADHD-friendly To-Do styling ===== */
.todo-section {
  border-radius: 16px;
  padding: 14px 14px;
  margin: 10px 0 14px 0;
  border: 1px solid rgba(0,0,0,0.10);
  box-shadow: 0 2px 10px rgba(0,0,0,0.06);
}
.todo-title {
  font-size: 1.05rem;
  font-weight: 800;
  letter-spacing: 0.2px;
  display:flex;
  align-items:center;
  gap:10px;
  margin-bottom: 10px;
}
.todo-title .count {
  font-size: 0.9rem;
  font-weight: 800;
  padding: 2px 10px;
  border-radius: 999px;
  border: 1px solid rgba(0,0,0,0.10);
  background: rgba(255,255,255,0.65);
}
.todo-card {
  border-radius: 14px;
  padding: 12px 12px;
  margin: 10px 0;
  border: 1px solid rgba(0,0,0,0.10);
  background: rgba(255,255,255,0.75);
}
.todo-card:hover {
  transform: translateY(-1px);
  box-shadow: 0 8px 18px rgba(0,0,0,0.10);
}
.todo-top {
  display:flex;
  justify-content:space-between;
  gap: 10px;
  align-items:flex-start;
  margin-bottom: 6px;
}
.todo-company {
  font-weight: 900;
  font-size: 1.02rem;
}
.todo-meta {
  font-size: 0.88rem;
  opacity: 0.78;
}
.badge {
  font-size: 0.80rem;
  font-weight: 900;
  padding: 3px 10px;
  border-radius: 999px;
  border: 1px solid rgba(0,0,0,0.10);
  white-space: nowrap;
}
.badge-overdue { background: rgba(255, 82, 82, 0.18); color: #7a0b0b; }
.badge-today   { background: rgba(255, 193, 7, 0.22); color: #5c3a00; }
.badge-soon    { background: rgba(76, 175, 80, 0.18); color: #0f4d1f; }

.strip {
  border-left: 6px solid rgba(0,0,0,0.12);
  padding-left: 10px;
}
.strip-overdue { border-left-color: rgba(255, 82, 82, 0.95); }
.strip-today   { border-left-color: rgba(255, 193, 7, 0.95); }
.strip-soon    { border-left-color: rgba(76, 175, 80, 0.95); }

.todo-action {
  display:flex;
  justify-content:flex-end;
  align-items:center;
  height: 100%;
}
</style>
""", unsafe_allow_html=True)

DB = "onboarding.db"

# ---------- Email Templates (IDs + subject + body) ----------
EMAIL_TEMPLATES = {
    "WLI-ONB-01": {
        "subject": "Understanding your setup",
        "body": """Hi {name},

Thank you for reaching out.

I am Ira, director at Wine Logistics International. Nice to meet you!

To make sure I provide the right rates and setup, could you briefly share:
- Storage requirements: approximate number of pallets and range of SKUs you’d expect to store
- Order volumes and frequency: estimated monthly inbound and outbound orders
- Transaction types: B2B sales, stock transfers, sample shipments, …?
- Packaging formats: e.g. 6x750ml, 12x750ml, …?
- Order destinations: within Belgium, across the EU, or also outside the EU?
- Product classification: are any products certified organic?

Once I have a bit more insight into your activities, I can provide rates and further info accordingly.

Best regards,
Ira
"""
    },
    "WLI-ONB-02": {
        "subject": "Rates & service overview",
        "body": """Hi {name},

Please find attached our general rate list and a basic cost simulation.
You can adjust the cells highlighted in green (e.g., number of months, number of cases) to estimate costs.

If this looks aligned, the next step is to formally set up your account and confirm the preferred customs setup.

Best regards,
Ira
"""
    },
    "WLI-ONB-03": {
        "subject": "Customer onboarding guide",
        "body": """Hi {name},

I’ve attached our Customer Onboarding Guide, which explains how storage, customs and deliveries work in practice.

For now, the only things we need from you are:
1) Confirmation of your preferred importer setup (winery as importer vs your customer as importer)
2) Whether you already have an EORI number (or if you’d like guidance to apply)

Happy to jump on a call if helpful.

Best regards,
Ira
"""
    },
    "WLI-ONB-04": {
        "subject": "Account setup – company details",
        "body": """Hi {name},

To set up your account with us and prepare the service agreement, could you please provide:

- Company name
- Legal form
- Registered office address
- Postal code and town/city
- Country
- Company registration number
- VAT identification number (if applicable)
- Name of the company representative
- Email address

Once we receive this information, I will prepare the service agreement for your review and signature.

Best regards,
Ira
"""
    },
    "WLI-ONB-05": {
        "subject": "Customs setup & compliance choices",
        "body": """Hi {name},

Based on the onboarding guide, we now need to confirm the following points:

1) Importer setup
- Will you act as the importer (typical for DDP), or will your EU customer act as the importer (typical for EXW/FCA/DAP)?

2) EORI
- Do you already have an EORI number? If not, we can guide you through the application (free and typically processed quickly).

3) Back-label importer
- For EU labeling, who should be listed as the importer on the back label?
  • Your EU customer (order-specific labels)
  • Your own EU entity (if applicable)
  • Wine Logistics International (universal EU label) – €750/year

4) Insurance
- Our coverage includes €25,000 per claim at no cost. Please confirm if this is sufficient or if you require a higher insured value.

Once we have these confirmations, we can finalize the agreements and proceed operationally.

Best regards,
Ira
"""
    },
    "WLI-ONB-06": {
        "subject": "Agreements & EORI",
        "body": """Hi {name},

Based on your selected setup, please find attached the required agreements for signature.
Once signed and (if applicable) once we have your EORI number, we are set to handle customs formalities on the agreed basis.

If you need guidance for the EORI application, let me know and we’ll assist.

Best regards,
Ira
"""
    },
    "WLI-ONB-07": {
        "subject": "Portal access & daily operations",
        "body": """Hi {name},

We will open an account for you and provide access to our online portal to check inventory and submit outgoing orders.

To set everything up, could you please confirm:
- The number of portal logins required + names and email addresses
- The contact(s) who should receive automatic departure notices for each order
- The contact(s) who should receive the monthly invoices

For daily operations, you can reach our team at orders@wine-logistics.com or +32 354 000 59.

Best regards,
Ira
"""
    },
    "WLI-ONB-08": {
        "subject": "Your WLI portal login",
        "body": """Hi {name},

Your login has been created.

Username: {username}
Password: {password}

Portal link:
https://erp.wine-logistics.com/login?ReturnUrl=%2f

The portal allows you to:
- View incoming shipments via “Infeeds”
- View prepared/departed orders via “Orders”
- Check inventory via “Stocklist”
- Trace product history via “ProductHistory”
- Track order status via “Order tracing”
- Submit outgoing orders via “Sending new order to WLI”
- Save drafts via “Saved orders”

Attached is a tutorial for submitting orders.

Best regards,
Ira
"""
    },

"WLI-ONB-09": {
    "subject": "Sales model, Incoterms & duties/taxes overview",
    "body": """Hi {name},

Before we proceed operationally, we would like to clarify the commercial and customs model for your shipments to the EU.

There are two main approaches:

1) Duty-paid model (DDP)
- You or your appointed importer handle import duties, VAT and excise upon entry into the EU
- This provides a smoother delivery experience for your customers
- Suitable for direct-to-customer or B2C/B2B deliveries

2) Duty-unpaid model (e.g. EXW / FCA / DAP)
- Your EU customer acts as the importer
- Duties, VAT and local excise taxes are settled by the importer in the destination country
- Often used for distributors or professional buyers

Key points to consider:
- Applicable Incoterms (DDP, DAP, FCA, EXW, etc.)
- Who is listed as the importer of record
- Invoice indications (value, Incoterm, importer details)
- EU import duties (where applicable)
- VAT at import
- Local excise duties and alcohol taxes depending on destination country

The chosen model has a direct impact on customs handling, labeling, documentation and invoicing structure.

We are happy to recommend the most suitable setup based on your order types, volumes and target markets.

Best regards,  
Ira
Wine Logistics International
"""
},

}

# ---------- Checklist definition (mapped to brochure + email IDs) ----------
# Excluding express/urgent scenarios as requested.
CHECKLIST_ITEMS = [
    # A. Prospect & commercial alignment
    ("A", "Introductory email or call", "§1–2", "WLI-ONB-01"),
    ("A", "Operational scope confirmed", "§1-2", "WLI-ONB-01"),
    ("A", "Rates sent", "§3–5", "WLI-ONB-02"),
    ("A", "Customs & compliance call", "§6-10", "WLI-ONB-05"),
    ("A", "Risk analysis", "Internal", None),

    # B. Account & legal setup
    ("B", "Company details received", "§11 / §13", "WLI-ONB-04"),
    ("B", "Service agreement sent", "—", "WLI-ONB-04"),
    ("B", "Service agreement signed", "—", "WLI-ONB-06"),
    ("B", "Account created in ERP", "Internal", None),
    ("B", "Account created in Odoo", "Internal", None),

    # C. Customs & compliance decisions
    ("C", "Sales models & Incoterms explained", "§6-10", "WLI-ONB-09"),
    ("C", "Sales model confirmed", "§6", "WLI-ONB-05"),
    ("C", "EORI confirmed / applied", "§9", "WLI-ONB-05"),

    ("C", "Customs representation agreement sent", "§6–7", "WLI-ONB-06"),
    ("C", "Fiscal representation agreement sent", "§6–7", "WLI-ONB-06"),

    ("C", "Customs representation agreement signed", "§6–7", "WLI-ONB-06"),
    ("C", "Fiscal representation agreement signed", "§6–7", "WLI-ONB-06"),

    ("C", "Back-label importer chosen", "§8", "WLI-ONB-05"),
    ("C", "Insurance level confirmed", "§10", "WLI-ONB-05"),

    # D. Operational setup
    ("D", "Operational contacts email sent", "§11", "WLI-ONB-07"),
    ("D", "Portal users defined", "§11", "WLI-ONB-07"),
    ("D", "Departure notification contacts set", "§11", "WLI-ONB-07"),
    ("D", "Invoice recipients set", "§11", "WLI-ONB-07"),
    ("D", "Order submission tutorial sent", "§11", "WLI-ONB-08"),

    # E. Internal handover
    ("E", "Orders team informed", "Internal", None),
    ("E", "Warehouse informed", "Internal", None),
    ("E", "Customer marked Operational", "—", "WLI-ONB-08"),
]

GROUP_LABELS = {
    "A": "A. Prospect & commercial alignment",
    "B": "B. Account & legal setup",
    "C": "C. Customs & compliance decisions",
    "D": "D. Operational setup",
    "E": "E. Internal handover",
}

# Status rules: what must be done to consider operational
REQUIRED_FOR_OPERATIONAL = [
    "Company details received",
    "Service agreement signed",
    "Importer model chosen",
    "EORI confirmed / applied",
    "Back-label importer chosen",
    "Insurance level confirmed",
    "Portal users defined",
    "Invoice recipients set",
    "Operational contacts email sent",
]

STATUS_CHOICES = ["New", "In progress", "On hold", "Operational", "Won", "Lost"]

QUICK_MAP = [
    ("SA", "Service agreement signed"),
    ("Customs", "Customs representation agreement signed"),
    ("Fiscal", "Fiscal representation agreement signed"),
    ("EORI", "EORI confirmed / applied"),
    ("Ins", "Insurance level confirmed"),
    ("Importer", "Back-label importer chosen"),
    ("WLI Portal", "__GROUP_D__"),
    ("Internal handover", "__GROUP_E__"),
]

def satisfied_from_rows(done: int, na: int) -> bool:
    return int(done) == 1 or int(na) == 1

def quick_status_for_company(df_check_all: pd.DataFrame, company_id: int) -> dict:
    dfc = df_check_all[df_check_all["company_id"] == company_id].copy()
    out = {}
    score = 0

    for short, item_name in QUICK_MAP:
        if item_name == "__GROUP_D__":
            ok = is_group_complete(dfc, "D")
        elif item_name == "__GROUP_E__":
            ok = is_group_complete(dfc, "E")
        else:
            row = dfc[dfc["item"] == item_name]
            ok = False
            if not row.empty:
                r = row.iloc[0]
                done = int(r.get("done", 0)) == 1
                na = int(r.get("na", 0)) == 1

                if done:
                    out[short] = "✅"
                elif na:
                    out[short] = "➖"
                else:
                    out[short] = "⬜"

        score += 1 if ok else 0

    out["Quick score"] = f"{score}/{len(QUICK_MAP)}"
    return out

def now_z():
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def conn():
    c = sqlite3.connect(DB, check_same_thread=False)
    c.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT,
            company_name TEXT,
            country TEXT,
            contact_name TEXT,
            contact_email TEXT,
            status TEXT,
            notes TEXT,
            next_followup_at TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS checklist (
            company_id INTEGER,
            grp TEXT,
            item TEXT,
            brochure_ref TEXT,
            email_ref TEXT,
            done INTEGER,
            na INTEGER DEFAULT 0,
            done_at TEXT,
            note TEXT,
            PRIMARY KEY(company_id, item)
        )
    """)

    c.execute("""
              CREATE TABLE IF NOT EXISTS company_events
              (
                  id
                  INTEGER
                  PRIMARY
                  KEY
                  AUTOINCREMENT,
                  company_id
                  INTEGER,

                  event_at
                  TEXT, -- ISO timestamp e.g. 2026-02-17T13:00:00Z
                  event_type
                  TEXT, -- Meeting, Warehouse visit, Call, Email, Task, Note
                  title
                  TEXT, -- short label
                  detail
                  TEXT, -- longer free text
                  created_at
                  TEXT, -- when the record was created

                  -- Task-specific fields (safe to keep for all event types)
                  is_done
                  INTEGER
                  DEFAULT
                  0,
                  done_at
                  TEXT,

                  -- Reminder fields (NULL = no reminder)
                  remind_minutes
                  INTEGER,
                  reminder_sent_at
                  TEXT
              )
              """)

    c.execute("""
              CREATE TABLE IF NOT EXISTS company_contacts
              (
                  id
                  INTEGER
                  PRIMARY
                  KEY
                  AUTOINCREMENT,
                  company_id
                  INTEGER,
                  name
                  TEXT,
                  email
                  TEXT,
                  role
                  TEXT,
                  is_primary
                  INTEGER
                  DEFAULT
                  0,
                  created_at
                  TEXT
              )
              """)

    return c


def init_company_checklist(c, company_id: int):
    cur = c.cursor()
    for idx, (grp, item, brochure_ref, email_ref) in enumerate(CHECKLIST_ITEMS):
        cur.execute("""
                    INSERT
                    OR IGNORE INTO checklist
            (company_id, grp, item, brochure_ref, email_ref, ord, done, done_at, note, na)
            VALUES (?, ?, ?, ?, ?, ?, 0, NULL, '', 0)
                    """, (company_id, grp, item, brochure_ref, email_ref, idx))
    c.commit()


def load_companies(c) -> pd.DataFrame:
    return pd.read_sql_query(
        "SELECT id, created_at, company_name, country, contact_name, contact_email, status FROM companies ORDER BY id DESC",
        c
    )


def load_company(c, company_id: int) -> pd.DataFrame:
    return pd.read_sql_query("SELECT * FROM companies WHERE id = ?", c, params=(company_id,))


def load_checklist(c, company_id):
    return pd.read_sql_query(
        """
        SELECT
            grp,
            item,
            brochure_ref,
            email_ref,
            done,
            na,
            done_at,
            note
        FROM checklist
        WHERE company_id = ?
        ORDER BY grp, COALESCE(ord, 9999), item
        """,
        c,
        params=(company_id,),
    )

def compute_next_email(df_check: pd.DataFrame):
    # Next email = the lowest-numbered WLI-ONB-0X among incomplete items
    pending = df_check[(df_check["done"] == 0) & (df_check["email_ref"].notna())]
    if pending.empty:
        return None
    # Sort by email id order
    def key(v):
        try:
            return int(v.split("-")[-1])
        except Exception:
            return 999
    email = sorted(pending["email_ref"].unique().tolist(), key=key)[0]
    # Provide the first pending item for context
    first_item = pending[pending["email_ref"] == email].iloc[0]["item"]
    return email, first_item


def can_set_operational(df_check: pd.DataFrame) -> (bool, list):
    done_items = set(df_check[df_check["done"] == 1]["item"].tolist())
    na_items = set(df_check[df_check["na"] == 1]["item"].tolist())
    satisfied = done_items | na_items
    missing = [x for x in REQUIRED_FOR_OPERATIONAL if x not in done_items and x not in na_items]
    return (len(missing) == 0, missing)


def generate_onboarding_excel(company_row, df_checklist):
    """
    company_row: a pandas Series (df_company.iloc[0])
    df_checklist: DataFrame from load_checklist(...)
    returns: BytesIO buffer ready for st.download_button
    """
    wb = Workbook()

    # ---------- Sheet 1: Overview ----------
    ws_overview = wb.active
    ws_overview.title = "Overview"

    total = len(df_checklist)
    done = int(df_checklist["done"].sum())
    progress = round((done / total) * 100, 1) if total else 0
    model = derive_sales_model(df_checklist)

    overview_rows = [
        ("Company name", company_row["company_name"]),
        ("Country", company_row["country"]),
        ("Contact", company_row["contact_name"]),
        ("Email", company_row["contact_email"]),
        ("Status", company_row["status"]),
        ("Sales model", f"{model['model']}"),
        ("Sales model note", model['hint']),
        ("Created at", fmt_date(company_row["created_at"], with_time=False)),
        ("Progress (%)", progress),
    ]

    for r, (k, v) in enumerate(overview_rows, start=1):
        ws_overview.cell(row=r, column=1, value=k)
        ws_overview.cell(row=r, column=2, value=v)

    ws_overview.column_dimensions["A"].width = 28
    ws_overview.column_dimensions["B"].width = 60

    # ---------- Quick status block on Overview ----------
    qs_df = quick_status_from_checklist(df_checklist)

    start_row = len(overview_rows) + 3  # leave 2 blank rows
    ws_overview.cell(row=start_row, column=1, value="Quick status")
    ws_overview.cell(row=start_row, column=1).font = ws_overview.cell(row=1, column=1).font.copy(bold=True)

    # Header
    ws_overview.cell(row=start_row + 1, column=1, value="Label")
    ws_overview.cell(row=start_row + 1, column=2, value="Status")

    # Rows
    r0 = start_row + 2
    for i, r in enumerate(qs_df.itertuples(index=False), start=0):
        ws_overview.cell(row=r0 + i, column=1, value=r.Label)
        ws_overview.cell(row=r0 + i, column=2, value=r.Status)

    # Make it readable
    ws_overview.column_dimensions["A"].width = 28
    ws_overview.column_dimensions["B"].width = 60


    # ---------- Sheet 2: Checklist ----------
    ws_check = wb.create_sheet(title="Checklist")

    headers = [
        "Group",
        "Group name",
        "Checklist item",
        "Done",
        "Done at",
        "Notes",
    ]
    ws_check.append(headers)

    for _, r in df_checklist.iterrows():
        ws_check.append([
            r["grp"],
            GROUP_LABELS.get(r["grp"], r["grp"]),
            r["item"],
            "Yes" if int(r["done"]) == 1 else "No",
            fmt_date(r["done_at"]) if r["done_at"] else "",
            r["note"] or "",
        ])

    # Make columns readable
    for col_idx in range(1, len(headers) + 1):
        ws_check.column_dimensions[get_column_letter(col_idx)].width = 22

    ws_check.column_dimensions["C"].width = 42  # Checklist item
    ws_check.column_dimensions["F"].width = 55  # Notes (was H before)

    # ---------- Output ----------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def load_dashboard_data(c):
    df_comp = pd.read_sql_query(
        "SELECT id, created_at, company_name, country, contact_name, contact_email, status, next_followup_at FROM companies",
        c
    )
    df_check = pd.read_sql_query(
        "SELECT company_id, grp, item, done, na, done_at, email_ref, note FROM checklist",
        c
    )
    return df_comp, df_check


def company_progress(df_check_company: pd.DataFrame) -> float:
    total = len(df_check_company)
    done = int(df_check_company["done"].sum()) if total else 0
    return round((done / total) * 100, 1) if total else 0.0


def latest_update_iso(df_check_company: pd.DataFrame):
    # done_at strings are ISO "....Z" -> lexicographic max works
    dates = df_check_company["done_at"].dropna().astype(str)
    return dates.max() if not dates.empty else None

AUTO_STATUSES = {"New", "In progress", "Operational"}
MANUAL_STATUSES = {"On hold", "Won", "Lost"}

START_SIGNALS = {
    "Prospect info received",
    "Rates sent",
    "Company details received",
}

def derive_status(current_status: str, df_check: pd.DataFrame) -> str:
    """
    Returns the recommended status based on checklist completion.
    Does NOT override manual statuses.
    """
    if current_status in MANUAL_STATUSES:
        return current_status

    done_items = set(df_check[df_check["done"] == 1]["item"].tolist())

    # Operational gate
    if all(req in done_items for req in REQUIRED_FOR_OPERATIONAL):
        return "Operational"

    # Started?
    if any(sig in done_items for sig in START_SIGNALS):
        return "In progress"

    return "New"


def backfill_new_checklist_items(c):
    cur = c.cursor()
    company_ids = [r[0] for r in cur.execute("SELECT id FROM companies").fetchall()]

    for company_id in company_ids:
        for idx, (grp, item, brochure_ref, email_ref) in enumerate(CHECKLIST_ITEMS):
            cur.execute("""
                        INSERT
                        OR IGNORE INTO checklist
                (company_id, grp, item, brochure_ref, email_ref, ord, done, done_at, note, na)
                VALUES (?, ?, ?, ?, ?, ?, 0, NULL, '', 0)
                        """, (company_id, grp, item, brochure_ref, email_ref, idx))

    c.commit()

def backfill_checklist_ord(c):
    """
    Sets checklist.ord based on the order of CHECKLIST_ITEMS.
    Does NOT touch done/done_at/na/note (so you keep all progress).
    """
    cur = c.cursor()

    # item -> order index based on the order in CHECKLIST_ITEMS
    order_map = {item: idx for idx, (grp, item, brochure_ref, email_ref) in enumerate(CHECKLIST_ITEMS)}

    # Update known items
    for item, idx in order_map.items():
        cur.execute("""
            UPDATE checklist
            SET ord = ?
            WHERE item = ?
        """, (idx, item))

    # Any rows that don't match CHECKLIST_ITEMS go to bottom
    cur.execute("""
        UPDATE checklist
        SET ord = 9999
        WHERE ord IS NULL
    """)

    c.commit()

def ensure_column(c, table, column, col_type):
    cols = [r[1] for r in c.execute(f"PRAGMA table_info({table})").fetchall()]
    if column not in cols:
        c.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
        c.commit()

def ensure_column_safe(c, table, column, col_type):
    cols = [r[1] for r in c.execute(f"PRAGMA table_info({table})").fetchall()]
    if column not in cols:
        c.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
        c.commit()

def build_timeline(df_company: pd.DataFrame, df_check: pd.DataFrame, df_events_custom: pd.DataFrame) -> pd.DataFrame:
    row = df_company.iloc[0]

    events = []

    # Company created
    events.append({
        "When": row["created_at"],
        "Type": "Created",
        "Detail": "Company record created",
        "Group": "",
        "Item": "",
        "Email ref": "",
        "Note": "",
    })

    # Checklist done events
    done_rows = df_check[df_check["done_at"].notna()].copy()
    for _, r in done_rows.iterrows():
        events.append({
            "When": r["done_at"],
            "Type": "Checklist – Done",
            "Detail": f"{r['item']}",
            "Group": r["grp"],
            "Item": r["item"],
            "Email ref": r["email_ref"] or "",
            "Note": r["note"] or "",
        })

    # ✅ ADD IT RIGHT HERE
    if df_events_custom is not None and not df_events_custom.empty:
        for _, e in df_events_custom.iterrows():
            title = e["title"] or e["event_type"] or "Event"
            detail = (e["detail"] or "").strip()

            is_task = (e.get("event_type") == "Task")
            is_done = int(e.get("is_done", 0)) == 1
            prefix = "✅ " if (is_task and is_done) else ("☐ " if is_task else "")

            events.append({
                "When": e["event_at"],
                "Type": f"Event – {e['event_type']}",
                "Detail": f"{prefix}{title}",
                "Group": "",
                "Item": "",
                "Email ref": "",
                "Note": detail,
            })

    # Then this continues unchanged
    df_events = pd.DataFrame(events)
    df_events = df_events.sort_values("When", ascending=True).reset_index(drop=True)
    return df_events


def build_vertical_timeline_html(df_events: pd.DataFrame) -> str:
    """
    Returns the exact HTML (CSS + body) for the vertical timeline.
    """
    if df_events is None or df_events.empty:
        return "<div style='padding:12px;'>No timeline events yet.</div>"

    css = """
    <style>
      .tl-wrap, .tl-wrap * {
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, "Noto Sans", "Apple Color Emoji", "Segoe UI Emoji", "Segoe UI Symbol", sans-serif;
        }
      .tl-wrap { margin: 0.5rem 0 1.5rem 0; }
      .tl-item { position: relative; display: flex; gap: 14px; padding: 0 0 18px 0; }
      .tl-rail { position: relative; width: 28px; flex: 0 0 28px; }
      .tl-line {
        position: absolute; left: 13px; top: 0; bottom: 0;
        width: 2px; background: rgba(0,0,0,0.15);
      }
      .tl-dot {
        position: relative; margin: 2px 0 0 6px;
        width: 14px; height: 14px; border-radius: 50%;
        background: white; border: 2px solid rgba(0,0,0,0.45);
        z-index: 2;
      }
      .tl-arrow {
        position: absolute; left: 9px; top: 26px;
        width: 0; height: 0;
        border-left: 5px solid transparent;
        border-right: 5px solid transparent;
        border-top: 7px solid rgba(0,0,0,0.25);
        z-index: 1;
      }
      .tl-card {
        flex: 1;
        padding: 10px 12px;
        border: 1px solid rgba(0,0,0,0.12);
        border-radius: 12px;
        box-shadow: 0 1px 2px rgba(0,0,0,0.06);
        background: rgba(255,255,255,0.65);
      }
      .tl-top {
        display: flex; flex-wrap: wrap;
        gap: 10px; align-items: baseline; justify-content: space-between;
        margin-bottom: 6px;
      }
      .tl-when { font-size: 0.85rem; opacity: 0.75; }
      .tl-type {
        font-size: 0.85rem;
        padding: 2px 8px;
        border-radius: 999px;
        border: 1px solid rgba(0,0,0,0.12);
        background: rgba(0,0,0,0.03);
        white-space: nowrap;
      }
      .tl-detail { font-size: 1rem; font-weight: 600; margin: 0 0 4px 0; }
      .tl-meta { font-size: 0.85rem; opacity: 0.85; }
      .tl-meta span { margin-right: 10px; }
      .tl-note { margin-top: 8px; padding: 6px 10px; font-size: 0.9rem; border-left: 4px solid #4f8bf9; background: rgba(79, 139, 249, 0.08); color: #1f3a8a; border-radius: 6px; }
      .tl-item:last-child { padding-bottom: 0; }
      .tl-item:last-child .tl-line { bottom: 10px; }
      .tl-item:last-child .tl-arrow { display: none; }

      /* Manual events (pink highlight) */
      .tl-manual .tl-dot { border-color: #e91e63; background: #fce4ec; }
      .tl-manual .tl-card { border-color: #f8bbd0; background: #fff0f6; }
      .tl-manual .tl-type { background: #f8bbd0; border-color: #f48fb1; color: #880e4f; }
      .tl-manual .tl-note { border-left: 4px solid #e91e63; background: #fde2f0; color: #880e4f; }

      /* 🟢 TASK EVENTS (green) */
      .tl-task .tl-dot { border-color: #2e7d32; background: #e8f5e9; }
      .tl-task .tl-card { border-color: #a5d6a7; background: #f1f8f4; }
      .tl-task .tl-type { background: #c8e6c9; border-color: #81c784; color: #1b5e20; }

      /* ✅ COMPLETED TASK (strong green) */
      .tl-task-done .tl-dot { border-color: #1b5e20; background: #c8e6c9; }
      .tl-task-done .tl-card { border-color: #66bb6a; background: #e8f5e9; opacity: 0.9; }
      .tl-task-done .tl-type { background: #66bb6a; border-color: #43a047; color: white; }
    </style>
    """

    items_html = []
    for _, e in df_events.iterrows():
        when = html.escape(fmt_date(str(e.get("When", ""))))
        typ = html.escape(str(e.get("Type", "")))
        typ_raw = str(e.get("Type", ""))

        is_task = "Event – Task" in typ_raw
        is_done_task = "✅" in str(e.get("Detail", ""))

        is_manual = typ_raw.startswith("Event –") and (not is_task)

        manual_class = " tl-manual" if is_manual else ""
        task_class = " tl-task" if is_task else ""
        done_task_class = " tl-task-done" if is_done_task else ""

        row_class = manual_class + task_class + done_task_class

        detail = html.escape(str(e.get("Detail", "")))
        email_ref = html.escape(str(e.get("Email ref", "")) if e.get("Email ref") else "")
        note = html.escape(str(e.get("Note", "")) if e.get("Note") else "")

        meta_bits = []
        if email_ref:
            meta_bits.append(f"<span><b>Email:</b> {email_ref}</span>")
        meta_html = f"<div class='tl-meta'>{''.join(meta_bits)}</div>" if meta_bits else ""

        note_html = f"<div class='tl-note'><b>Note:</b> {note}</div>" if note else ""

        items_html.append(f"""
          <div class="tl-item{row_class}">
            <div class="tl-rail">
              <div class="tl-line"></div>
              <div class="tl-dot"></div>
              <div class="tl-arrow"></div>
            </div>
            <div class="tl-card">
              <div class="tl-top">
                <div class="tl-when">{when}</div>
                <div class="tl-type">{typ}</div>
              </div>
              <div class="tl-detail">{detail}</div>
              {meta_html}
              {note_html}
            </div>
          </div>
        """)

    return css + "<div class='tl-wrap'>" + "".join(items_html) + "</div>"

def render_vertical_timeline(df_events: pd.DataFrame):
    """
    Renders the vertical timeline in Streamlit using the same HTML builder
    we will also use for the PDF export.
    """
    timeline_html = build_vertical_timeline_html(df_events)

    # keep your old behavior for empty
    if df_events is None or df_events.empty:
        st.info("No timeline events yet.")
        return

    components.html(timeline_html, height=800, scrolling=True)

def html_to_pdf_bytes(html_body: str) -> BytesIO:
    """
    Renders html_body in headless Chromium and prints to PDF.
    Returns BytesIO.
    """
    # Wrap your body in a full HTML document
    full_html = f"""
    <!doctype html>
    <html>
      <head>
        <meta charset="utf-8"/>
        <meta name="viewport" content="width=device-width, initial-scale=1"/>
      </head>
      <body>
        {html_body}
      </body>
    </html>
    """

    buf = BytesIO()

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(viewport={"width": 1200, "height": 800})

        # Load the HTML
        page.set_content(full_html, wait_until="networkidle")

        # IMPORTANT: print full length; if your timeline is long, Chromium will paginate
        pdf_bytes = page.pdf(
            format="A4",
            print_background=True,
            margin={"top": "12mm", "bottom": "12mm", "left": "12mm", "right": "12mm"},
        )

        browser.close()

    buf.write(pdf_bytes)
    buf.seek(0)
    return buf

def fmt_date(iso_ts: str, with_time: bool = True) -> str:
    """
    Convert ISO timestamp (YYYY-MM-DDTHH:MM:SSZ)
    to DD-MM-YYYY (optionally with time).
    """
    if not iso_ts:
        return "—"
    try:
        dt = datetime.fromisoformat(iso_ts.replace("Z", ""))
        return dt.strftime("%d-%m-%Y %H:%M") if with_time else dt.strftime("%d-%m-%Y")
    except Exception:
        return iso_ts

from datetime import timedelta

def parse_iso_z(s: str):
    if not s:
        return None
    try:
        return datetime.fromisoformat(str(s).replace("Z", ""))
    except Exception:
        return None

def days_since(iso_ts: str):
    dt = parse_iso_z(iso_ts)
    if not dt:
        return None
    return (datetime.utcnow() - dt).days

def is_due_today(iso_ts: str):
    dt = parse_iso_z(iso_ts)
    if not dt:
        return False
    return dt.date() == datetime.utcnow().date()

def is_overdue(iso_ts: str):
    dt = parse_iso_z(iso_ts)
    if not dt:
        return False
    return dt.date() < datetime.utcnow().date()

def calc_progress(df_check_company: pd.DataFrame) -> float:
    applicable = df_check_company[df_check_company["na"] == 0]
    total = len(applicable)
    done = int(applicable["done"].sum()) if total else 0
    return round((done / total) * 100, 1) if total else 0.0

def add_event(c, company_id: int, event_at_iso: str, event_type: str, title: str, detail: str,
              is_done: int = 0, done_at: str | None = None,
              remind_minutes: int | None = None):
    c.execute("""
        INSERT INTO company_events (company_id, event_at, event_type, title, detail, created_at, is_done, done_at, remind_minutes, reminder_sent_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, NULL)
    """, (company_id, event_at_iso, event_type, title, detail, now_z(), int(is_done), done_at, remind_minutes))
    c.commit()

def load_events(c, company_id: int) -> pd.DataFrame:
    return pd.read_sql_query("""
        SELECT id, company_id, event_at, event_type, title, detail, created_at, is_done, done_at
        FROM company_events
        WHERE company_id = ?
        ORDER BY event_at ASC
    """, c, params=(company_id,))

def update_event(c, event_id: int, event_at_iso: str, event_type: str, title: str, detail: str):
    c.execute("""
        UPDATE company_events
        SET event_at = ?, event_type = ?, title = ?, detail = ?
        WHERE id = ?
    """, (event_at_iso, event_type, title, detail, event_id))
    c.commit()

def delete_event(c, event_id: int):
    c.execute("DELETE FROM company_events WHERE id = ?", (event_id,))
    c.commit()

def toggle_task_done(event_id: int, key: str):
    done_val = bool(st.session_state.get(key, False))
    set_task_done(c, event_id, done_val)
    st.toast("Task marked done ✅" if done_val else "Task reopened", icon="✅" if done_val else "↩️")

def set_task_done(c, event_id: int, is_done: bool):
    """
    Marks a Task event as done/undone.
    - is_done=True  -> sets is_done=1 and done_at=now
    - is_done=False -> sets is_done=0 and clears done_at
    """
    if is_done:
        c.execute("""
            UPDATE company_events
            SET is_done = 1,
                done_at = ?
            WHERE id = ?
        """, (now_z(), int(event_id)))
    else:
        c.execute("""
            UPDATE company_events
            SET is_done = 0,
                done_at = NULL
            WHERE id = ?
        """, (int(event_id),))
    c.commit()

def add_contact(c, company_id: int, name: str, email: str, role: str = "", is_primary: int = 0):
    c.execute("""
        INSERT INTO company_contacts (company_id, name, email, role, is_primary, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (company_id, name.strip(), email.strip(), role.strip(), int(is_primary), now_z()))
    c.commit()

def load_contacts(c, company_id: int) -> pd.DataFrame:
    return pd.read_sql_query("""
        SELECT id, name, email, role, is_primary
        FROM company_contacts
        WHERE company_id = ?
        ORDER BY is_primary DESC, id ASC
    """, c, params=(company_id,))

def iso_from_date_time(d, t):
    """Convert date + time to your ISO '...Z' format."""
    return datetime(d.year, d.month, d.day, t.hour, t.minute, 0).strftime("%Y-%m-%dT%H:%M:%SZ")

def is_item_satisfied(df_check: pd.DataFrame, item_name: str) -> bool:
    row = df_check[df_check["item"] == item_name]
    if row.empty:
        return False
    r = row.iloc[0]
    return (int(r.get("done", 0)) == 1) or (int(r.get("na", 0)) == 1)

def last_activity_for_all_companies(c) -> pd.DataFrame:
    """
    Returns DataFrame with columns: company_id, last_check_done_at, last_event_at
    """
    df_last_check = pd.read_sql_query("""
        SELECT company_id, MAX(done_at) AS last_check_done_at
        FROM checklist
        WHERE done_at IS NOT NULL
        GROUP BY company_id
    """, c)

    df_last_event = pd.read_sql_query("""
        SELECT company_id, MAX(event_at) AS last_event_at
        FROM company_events
        WHERE event_at IS NOT NULL
        GROUP BY company_id
    """, c)

    df = df_last_check.merge(df_last_event, on="company_id", how="outer")
    return df

def quick_status_from_checklist(df_checklist: pd.DataFrame) -> pd.DataFrame:
    """
    Returns a small DataFrame with columns: Key, Label, Status
    Includes:
    - Item-based quick status (SA, EORI, etc.)
    - Group-based status:
        D = WLI portal (Operational setup)
        E = Internal handover
    """

    def is_item_satisfied(item_name: str) -> bool:
        row = df_checklist[df_checklist["item"] == item_name]
        if row.empty:
            return False
        r = row.iloc[0]
        return (int(r.get("done", 0)) == 1) or (int(r.get("na", 0)) == 1)

    def is_group_complete(group_letter: str) -> bool:
        group_rows = df_checklist[df_checklist["grp"] == group_letter]

        if group_rows.empty:
            return False

        # Only count applicable items (not N/A)
        applicable = group_rows[group_rows["na"] == 0]

        if applicable.empty:
            # If everything is N/A, consider it complete
            return True

        # All applicable items must be done
        return bool((applicable["done"].astype(int) == 1).all())

    quick_map = [
        ("SA", "Service agreement", ("item", "Service agreement signed")),
        ("Customs", "Customs rep. agreement", ("item", "Customs representation agreement signed")),
        ("Fiscal", "Fiscal rep. agreement", ("item", "Fiscal representation agreement signed")),
        ("EORI", "EORI", ("item", "EORI confirmed / applied")),
        ("Ins", "Insurance", ("item", "Insurance level confirmed")),
        ("Importer", "Importer details", ("item", "Back-label importer chosen")),

        # NEW GROUP-BASED STATUS
        ("Portal", "WLI portal", ("group", "D")),
        ("Handover", "Internal handover", ("group", "E")),
    ]

    rows = []
    score = 0

    for key, label, rule in quick_map:
        rule_type, value = rule

        if rule_type == "item":
            ok = is_item_satisfied(value)
        else:  # group
            ok = is_group_complete(value)

        score += 1 if ok else 0

        rows.append({
            "Key": key,
            "Label": label,
            "Status": "✅" if ok else "⬜"
        })

    rows.append({
        "Key": "",
        "Label": "Quick score",
        "Status": f"{score}/{len(quick_map)}"
    })

    return pd.DataFrame(rows)

def item_state(df_check: pd.DataFrame, item_name: str) -> str:
    """
    Returns: 'done' | 'na' | 'open'
    """
    row = df_check[df_check["item"] == item_name]
    if row.empty:
        return "open"
    r = row.iloc[0]
    if int(r.get("done", 0)) == 1:
        return "done"
    if int(r.get("na", 0)) == 1:
        return "na"
    return "open"


def derive_sales_model(df_check: pd.DataFrame) -> dict:
    """
    Business rule you described:
    - If customs + fiscal agreements are N/A => Duty-unpaid (customer is importer)
    - If they are checked/signed => Duty-paid (WLI/importer setup)
    - Otherwise => Pending/Unknown
    """
    customs = item_state(df_check, "Customs representation agreement signed")
    fiscal = item_state(df_check, "Fiscal representation agreement signed")

    # duty-unpaid when both are explicitly N/A
    if customs == "na" and fiscal == "na":
        return {
            "model": "Duty-unpaid",
            "hint": "Customs & fiscal representation are N/A → customer is importer (typical duty-unpaid setup).",
            "icon": "🟦",
        }

    # duty-paid when either is signed (or both signed)
    if customs == "done" or fiscal == "done":
        return {
            "model": "Duty-paid",
            "hint": "Customs/fiscal representation signed → duty-paid/importer service is in place.",
            "icon": "🟩",
        }

    # otherwise not decided yet
    return {
        "model": "Pending",
        "hint": "Sales model not fully confirmed yet (customs/fiscal representation not decided).",
        "icon": "🟨",
    }

def update_company(c, company_id: int, company_name: str, country: str,
                   contact_name: str, contact_email: str, status: str, notes: str):
    c.execute("""
        UPDATE companies
        SET company_name = ?,
            country = ?,
            contact_name = ?,
            contact_email = ?,
            status = ?,
            notes = ?
        WHERE id = ?
    """, (company_name, country, contact_name, contact_email, status, notes, company_id))
    c.commit()

def delete_company(c, company_id: int):
    # Important: delete child rows first (no FK cascade in your schema)
    c.execute("DELETE FROM checklist WHERE company_id = ?", (company_id,))
    c.execute("DELETE FROM company_events WHERE company_id = ?", (company_id,))
    c.execute("DELETE FROM company_contacts WHERE company_id = ?", (company_id,))
    c.execute("DELETE FROM companies WHERE id = ?", (company_id,))
    c.commit()

def is_row_satisfied(r) -> bool:
    return int(r.get("done", 0)) == 1 or int(r.get("na", 0)) == 1

def is_group_complete(df_check: pd.DataFrame, grp: str) -> bool:
    g = df_check[df_check["grp"] == grp].copy()
    if g.empty:
        return False
    # Only count applicable items (na==0) toward “must be done”
    applicable = g[g["na"] == 0]
    if applicable.empty:
        # If everything is N/A, treat as complete
        return True
    return bool((applicable["done"].astype(int) == 1).all())

def load_all_tasks(c) -> pd.DataFrame:
    return pd.read_sql_query("""
        SELECT e.id, e.company_id, e.event_at, e.title, e.detail, e.is_done,
               co.company_name, co.status
        FROM company_events e
        JOIN companies co ON co.id = e.company_id
        WHERE e.event_type = 'Task'
    """, c)

def update_task_detail(c, event_id: int, title: str, detail: str):
    c.execute("""
        UPDATE company_events
        SET title = ?, detail = ?
        WHERE id = ?
    """, (title, detail, event_id))
    c.commit()

def render_month_calendar_counts(df_companies: pd.DataFrame, date_col: str, title: str = "New prospects"):
    """
    Calendar grid for counts per day in a selected month.
    df_companies[date_col] must contain ISO strings like 2026-02-17T10:22:00Z (or None).
    """
    # Parse to datetime (UTC-ish; you store Z)
    dts = pd.to_datetime(df_companies[date_col], errors="coerce").dt.date
    df = df_companies.copy()
    df["_d"] = dts
    df = df.dropna(subset=["_d"])

    # Month picker default: current month
    today = datetime.utcnow().date()
    ym = st.selectbox(
        "Month",
        options=[(y, m) for y in range(today.year - 2, today.year + 1) for m in range(1, 13)],
        index=[(y, m) for y in range(today.year - 2, today.year + 1) for m in range(1, 13)].index((today.year, today.month)),
        format_func=lambda x: f"{calendar.month_name[x[1]]} {x[0]}",
        key=f"cal_month_{date_col}",
    )
    year, month = ym

    # Filter to that month
    first = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    last = date(year, month, last_day)

    dfm = df[(df["_d"] >= first) & (df["_d"] <= last)].copy()

    # Count per day
    counts = dfm.groupby("_d").size().to_dict()

    # Header
    st.subheader(f"{title} — {calendar.month_name[month]} {year}")
    st.caption(f"Total this month: **{len(dfm)}**")

    # Calendar grid (Mon..Sun)
    cal = calendar.Calendar(firstweekday=0)  # Monday
    weeks = cal.monthdatescalendar(year, month)

    # Determine intensity scaling
    max_count = max(counts.values()) if counts else 0

    # Render grid with columns
    dow = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    cols = st.columns(7)
    for i, d in enumerate(dow):
        cols[i].markdown(f"**{d}**")

    for w in weeks:
        cols = st.columns(7)
        for i, day in enumerate(w):
            in_month = (day.month == month)
            c = counts.get(day, 0)

            # simple shading using emoji blocks
            if not in_month:
                cols[i].markdown("<div style='opacity:0.35'>—</div>", unsafe_allow_html=True)
                continue

            if c == 0:
                shade = "⬜"
            else:
                # scale 1..4 blocks
                level = 1 if max_count == 0 else max(1, min(4, int(round((c / max_count) * 4))))
                shade = {1:"🟩", 2:"🟩🟩", 3:"🟩🟩🟩", 4:"🟩🟩🟩🟩"}[level]

            cols[i].markdown(
                f"""
                <div style="
                    border:1px solid rgba(0,0,0,0.12);
                    border-radius:10px;
                    padding:8px;
                    min-height:64px;
                ">
                    <div style="font-size:12px; opacity:0.75;">{day.day:02d}</div>
                    <div style="font-size:18px; font-weight:700;">{c}</div>
                    <div style="margin-top:4px;">{shade}</div>
                </div>
                """,
                unsafe_allow_html=True
            )

def rename_checklist_item(c, old_name: str, new_name: str):
    c.execute("UPDATE checklist SET item = ? WHERE item = ?", (new_name, old_name))
    c.commit()

def keep_group_open(company_id: int, grp: str):
    st.session_state[f"exp_open_{company_id}_{grp}"] = True


c = conn()
ensure_column(c, "checklist", "na", "INTEGER DEFAULT 0")
ensure_column(c, "checklist", "ord", "INTEGER")
ensure_column(c, "companies", "next_followup_at", "TEXT")

ensure_column(c, "company_events", "is_done", "INTEGER DEFAULT 0")
ensure_column(c, "company_events", "done_at", "TEXT")
ensure_column(c, "company_events", "remind_minutes", "INTEGER")   # <-- better
ensure_column(c, "company_events", "reminder_sent_at", "TEXT")

# --- Checklist item renames (to preserve history) ---
rename_checklist_item(c, "Introductory email", "Introductory email or call")
rename_checklist_item(c, "Prospect info received", "Operational scope confirmed")
rename_checklist_item(c, "Importer model chosen", "Sales model confirmed")

backfill_new_checklist_items(c)
backfill_checklist_ord(c)

st.title("WLI Onboarding")

PAGES = [
    "📊 Today",
    "📅 Calendar / Analytics",
    "➕ New company",
    "📇 Companies",
    "🏢 Company record",
    "📋 Pipeline / Checklist",
    "📧 Email library",
    "🕘 Timeline"
]

with st.sidebar:
    st.header("Navigation")

    # Safety: if nav_page somehow isn't in PAGES, reset it
    if st.session_state.nav_page not in PAGES:
        st.session_state.nav_page = PAGES[0]

    st.radio("Go to", PAGES, key="nav_page")

page = st.session_state.nav_page

df_comp, df_check_all = load_dashboard_data(c)

# Consider "open" = not Won/Lost
open_mask = ~df_comp["status"].isin(["Won", "Lost"])
open_companies = df_comp[open_mask].copy()

open_count = len(open_companies)
new_count = int((open_companies["status"] == "New").sum())
inprog_count = int((open_companies["status"] == "In progress").sum())
onhold_count = int((open_companies["status"] == "On hold").sum())

# At-risk definition (simple): On hold OR no checklist updates yet while status is In progress
# (You can refine this later to "no update in X days")
at_risk_ids = set(open_companies[open_companies["status"] == "On hold"]["id"].tolist())

# KPI row
if page == "📊 Today":
    # KPI row
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Open onboardings", open_count)
    k2.metric("New", new_count)
    k3.metric("In progress", inprog_count)
    k4.metric("On hold", onhold_count)

    st.divider()

    st.subheader("To-do list (all customers)")

    df_tasks_all = load_all_tasks(c)

    # only open pipeline customers
    df_tasks_all = df_tasks_all[~df_tasks_all["status"].isin(["Won", "Lost"])].copy()

    # open tasks only
    df_open_tasks = df_tasks_all[df_tasks_all["is_done"].astype(int) == 0].copy()

    if df_open_tasks.empty:
        st.success("No open tasks 🎉")
    else:
        # Helpers for sorting + bucketing
        df_open_tasks["dt"] = df_open_tasks["event_at"].apply(parse_iso_z)
        df_open_tasks = df_open_tasks[df_open_tasks["dt"].notna()].copy()
        df_open_tasks["overdue"] = df_open_tasks["event_at"].apply(is_overdue)
        df_open_tasks["due_today"] = df_open_tasks["event_at"].apply(is_due_today)
        df_open_tasks["days_until"] = df_open_tasks["dt"].apply(lambda x: (x.date() - datetime.utcnow().date()).days)

        # Optional filters
        show_upcoming_days = st.slider("Show upcoming (days)", 1, 30, 14, 1, key="todo_days")
        only_company = st.selectbox(
            "Filter by company (optional)",
            ["All"] + sorted(df_open_tasks["company_name"].unique().tolist()),
            key="todo_company"
        )
        if only_company != "All":
            df_open_tasks = df_open_tasks[df_open_tasks["company_name"] == only_company].copy()

        # Buckets
        df_overdue = df_open_tasks[df_open_tasks["overdue"] == True].sort_values("dt")
        df_today = df_open_tasks[df_open_tasks["due_today"] == True].sort_values("dt")
        df_upcoming = df_open_tasks[
            (df_open_tasks["overdue"] == False)
            & (df_open_tasks["due_today"] == False)
            & (df_open_tasks["days_until"] <= show_upcoming_days)
            & (df_open_tasks["days_until"] >= 0)
            ].sort_values("dt")


        def render_task_bucket(df_bucket: pd.DataFrame, label: str, kind: str):
            """
            kind: 'overdue' | 'today' | 'soon' (controls color/badge)
            """
            if df_bucket.empty:
                return

            icon = {"overdue": "⛔", "today": "⚡", "soon": "🟢"}[kind]
            badge_class = {"overdue": "badge-overdue", "today": "badge-today", "soon": "badge-soon"}[kind]
            strip_class = {"overdue": "strip-overdue", "today": "strip-today", "soon": "strip-soon"}[kind]

            st.markdown(
                f"""
                <div class="todo-section">
                  <div class="todo-title">{icon} {label} <span class="count">{len(df_bucket)}</span></div>
                </div>
                """,
                unsafe_allow_html=True
            )

            for _, t in df_bucket.iterrows():
                tid = int(t["id"])
                title = (t["title"] or "").strip() or "(no title)"
                detail = (t["detail"] or "").strip()

                # computed earlier in your code: t["days_until"] exists
                days_until = int(t.get("days_until", 0))

                if kind == "overdue":
                    badge = f'<span class="badge {badge_class}">OVERDUE</span>'
                    due_hint = "Due date passed"
                elif kind == "today":
                    badge = f'<span class="badge {badge_class}">TODAY</span>'
                    due_hint = "Due today"
                else:
                    badge = f'<span class="badge {badge_class}">IN {days_until}D</span>'
                    due_hint = f"Due in {days_until} day(s)"

                left, right = st.columns([6, 1.2])

                with left:
                    st.markdown(
                        f"""
                        <div class="todo-card strip {strip_class}">
                          <div class="todo-top">
                            <div>
                              <div class="todo-company">{html.escape(str(t["company_name"]))}</div>
                              <div class="todo-meta">
                                <b>{html.escape(title)}</b><br/>
                                {html.escape(due_hint)} • {html.escape(fmt_date(t["event_at"]))}
                              </div>
                            </div>
                            <div>{badge}</div>
                          </div>
                          {"<div class='todo-meta'>" + html.escape(detail) + "</div>" if detail else ""}
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                with right:
                    # Make the action really clear and consistent
                    done_now = st.button("✅ Done", key=f"todo_done_btn_{tid}", use_container_width=True)
                    if done_now:
                        set_task_done(c, tid, True)
                        st.toast("Task marked done ✅", icon="✅")
                        st.rerun()

            st.divider()


        focus_mode = st.checkbox("🎯 Focus mode (show only Overdue + Today)", value=True)
        if focus_mode:
            df_upcoming = df_upcoming.iloc[0:0]  # hide upcoming

        render_task_bucket(df_overdue, "OverDUE — do these first", "overdue")
        render_task_bucket(df_today, "Due TODAY — quick wins", "today")
        render_task_bucket(df_upcoming, f"Upcoming — next {show_upcoming_days} days", "soon")

    st.subheader("Quick status per customer")

    # Load data
    df_comp, df_check_all = load_dashboard_data(c)

    # Open = exclude Won/Lost
    open_mask = ~df_comp["status"].isin(["Won", "Lost"])
    df_open = df_comp[open_mask].copy()

    df_last = last_activity_for_all_companies(c)

    # Build quick status rows
    rows_q = []
    for _, comp in df_open.iterrows():
        cid = int(comp["id"])

        # --- quick status ---
        qs = quick_status_for_company(df_check_all, cid)

        # --- last activity (needs cid, so it MUST be inside loop) ---
        last_row = df_last[df_last["company_id"] == cid]

        last_check = None
        last_event = None

        if not last_row.empty:
            lc = last_row["last_check_done_at"].iloc[0]
            le = last_row["last_event_at"].iloc[0]

            if pd.notna(lc):
                last_check = str(lc)
            if pd.notna(le):
                last_event = str(le)

        dt_check = parse_iso_z(last_check) if last_check else None
        dt_event = parse_iso_z(last_event) if last_event else None

        if dt_check and dt_event:
            last_activity_dt = max(dt_check, dt_event)
        elif dt_check:
            last_activity_dt = dt_check
        elif dt_event:
            last_activity_dt = dt_event
        else:
            last_activity_dt = None

        last_activity = (
            last_activity_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
            if last_activity_dt else None
        )

        # --- append row ---
        rows_q.append({
            "Company": comp["company_name"],
            "Status": comp["status"],
            "Last activity": fmt_date(last_activity, with_time=False) if last_activity else "—",
            "Quick score": qs["Quick score"],
            "SA": qs["SA"],
            "Customs": qs["Customs"],
            "Fiscal": qs["Fiscal"],
            "EORI": qs["EORI"],
            "Ins": qs["Ins"],
            "Importer": qs["Importer"],
        })

    df_quick = pd.DataFrame(rows_q)

    # Filter toggle (ONLY after df_quick exists)
    only_incomplete = st.checkbox(
        f"Show only customers not fully ready (Quick status < {len(QUICK_MAP)}/{len(QUICK_MAP)})",
        value=True,
        key="today_only_incomplete_quick"
    )

    if not df_quick.empty:
        if only_incomplete:
            df_quick = df_quick[
                df_quick["Quick score"] != f"{len(QUICK_MAP)}/{len(QUICK_MAP)}"
            ]

        # Sort by lowest completion first (most urgent on top)
        def score_num(s):
            try:
                return int(str(s).split("/")[0])
            except Exception:
                return 0

        df_quick["__score"] = df_quick["Quick score"].apply(score_num)
        df_quick = df_quick.sort_values(
            by=["__score", "Status", "Company"],
            ascending=[True, True, True]
        ).drop(columns="__score")

        # 🔥 ONLY ONE TABLE (the one you actually want)
        st.dataframe(df_quick, use_container_width=True, hide_index=True)
    else:
        st.info("No open customers yet.")

    st.subheader("Follow-ups due")

    df_comp, _ = load_dashboard_data(c)
    df_open = df_comp[~df_comp["status"].isin(["Won", "Lost"])].copy()

    df_open["overdue_fu"] = df_open["next_followup_at"].apply(is_overdue)
    df_open["due_today_fu"] = df_open["next_followup_at"].apply(is_due_today)

    df_fu = df_open[(df_open["overdue_fu"] == True) | (df_open["due_today_fu"] == True)].copy()

    if df_fu.empty:
        st.caption("No follow-ups due.")
    else:
        df_fu["Due"] = df_fu["next_followup_at"].apply(lambda x: fmt_date(x))
        st.dataframe(
            df_fu[["company_name", "status", "Due"]].rename(columns={"company_name": "Company", "status": "Status"}),
            use_container_width=True,
            hide_index=True
        )

# ------------------- TAB 1 -------------------
if page == "➕ New company":
    st.subheader("Create a new company record")

    # Session state: number of contact blocks
    if "new_company_contact_count" not in st.session_state:
        st.session_state.new_company_contact_count = 1

    with st.form("new_company"):
        company_name = st.text_input("Company name *")
        country = st.text_input("Country *")

        contacts = []
        for i in range(st.session_state.new_company_contact_count):
            st.markdown(f"Contact {i+1}")
            c1, c2, c3 = st.columns([2, 2, 1.5])
            with c1:
                name = st.text_input("Name *" if i == 0 else "Name", key=f"nc_name_{i}")
            with c2:
                email = st.text_input("Email *" if i == 0 else "Email", key=f"nc_email_{i}")
            with c3:
                role = st.text_input("Role", placeholder="e.g. Logistics / Finance", key=f"nc_role_{i}")

            # store (allow blanks for extra contacts; validate later)
            contacts.append({"name": name, "email": email, "role": role})

        # Add contact button (inside form = needs form_submit_button trick)
        add_more = st.form_submit_button("➕ Add another contact")

        status = st.selectbox("Status", STATUS_CHOICES, index=0)
        notes = st.text_area("Notes")

        submit = st.form_submit_button("Create")

    # Handle add contact click (rerun-safe)
    if add_more:
        st.session_state.new_company_contact_count += 1
        st.rerun()

    if submit:
        # Basic company validation
        missing = []
        if not company_name.strip(): missing.append("Company name")
        if not country.strip(): missing.append("Country")

        # Require at least 1 complete contact
        primary = contacts[0]
        if not primary["name"].strip(): missing.append("Primary contact name")
        if not primary["email"].strip(): missing.append("Primary contact email")

        if missing:
            st.error("Missing required fields: " + ", ".join(missing))
        else:
            # ---------- DUPLICATE CHECK (ADD THIS BLOCK HERE) ----------
            company_name_clean = company_name.strip()
            email_clean = primary["email"].strip()

            possible_dupes = pd.read_sql_query("""
                SELECT id, company_name, contact_email, country
                FROM companies
                WHERE LOWER(company_name) = LOWER(?)
                   OR LOWER(contact_email) = LOWER(?)
                ORDER BY id DESC
            """, c, params=(company_name_clean, email_clean))

            if not possible_dupes.empty:
                st.error("⚠️ Possible duplicate customer detected!")
                st.write("Existing record(s):")
                st.dataframe(possible_dupes, use_container_width=True, hide_index=True)

                st.info("If this is intentional, slightly adjust the name (e.g. add region) or delete the duplicate first.")
                st.stop()
            # ---------- END DUPLICATE CHECK ----------

            # Create company (keep primary contact in companies table for compatibility)
            cur = c.cursor()
            cur.execute("""
                INSERT INTO companies (created_at, company_name, country, contact_name, contact_email, status, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                now_z(),
                company_name_clean,
                country.strip(),
                primary["name"].strip(),
                email_clean,
                status,
                notes.strip()
            ))
            company_id = cur.lastrowid
            c.commit()

            init_company_checklist(c, company_id)

            # Save contacts table: include primary + any additional non-empty contacts
            add_contact(c, company_id, primary["name"], email_clean, primary.get("role",""), is_primary=1)

            for extra in contacts[1:]:
                if extra["name"].strip() or extra["email"].strip():
                    # require both if any is filled
                    if not extra["name"].strip() or not extra["email"].strip():
                        st.warning("Extra contacts need both name + email. Skipped an incomplete contact.")
                        continue
                    add_contact(c, company_id, extra["name"], extra["email"], extra.get("role",""), is_primary=0)

            st.success(f"Created {company_name} ✅ (ID {company_id})")

            # reset contact count for next entry
            st.session_state.new_company_contact_count = 1



# ------------------- TAB 2 -------------------
if page == "📋 Pipeline / Checklist":
    st.subheader("Pipeline")
    df_companies = load_companies(c)

    if df_companies.empty:
        st.info("No companies yet. Add one in the 'New company' tab.")
    else:
        # ✅ Start with select a company
        options = df_companies[["id", "company_name"]].values.tolist()
        # ✅ Calendar -> Pipeline jump support
        preselect_id = st.session_state.get("pipeline_open_company_id", None)

        options = df_companies[["id", "company_name"]].values.tolist()

        default_index = 0
        if preselect_id is not None:
            for idx, opt in enumerate(options):
                if int(opt[0]) == int(preselect_id):
                    default_index = idx
                    break

        selected = st.selectbox(
            "Select a company",
            options,
            index=default_index,
            format_func=lambda x: f"{x[1]} (ID {x[0]})"
        )

        company_id = int(selected[0])

        # Optional: clear the preselect so next time it doesn’t keep forcing the same company
        if "pipeline_open_company_id" in st.session_state:
            del st.session_state.pipeline_open_company_id

        df_company = load_company(c, company_id)
        df_check = load_checklist(c, company_id)

        # Header: status + next email suggestion + progress
        left, mid, right = st.columns([2, 2, 2])

        with left:
            row = df_company.iloc[0]

            # Load contacts from new table
            df_contacts = load_contacts(c, company_id)

            if df_contacts.empty:
                # Fallback to legacy single-contact fields
                st.write(f"**Contact:** {row['contact_name']} — {row['contact_email']}")
            else:
                st.write("**Contacts:**")
                for _, r in df_contacts.iterrows():
                    tag = " (primary)" if int(r["is_primary"]) == 1 else ""
                    role = f" — {r['role']}" if str(r.get("role", "")).strip() else ""
                    st.write(f"- {r['name']} — {r['email']}{role}{tag}")

            st.write(f"**Country:** {row['country']}")
            st.write(f"**Created:** {fmt_date(row['created_at'], with_time=False)}")

            excel_buffer = generate_onboarding_excel(row, df_check)

            safe_name = "".join(ch for ch in row["company_name"] if ch.isalnum() or ch in (" ", "-", "_")).strip()
            file_name = f"WLI_Onboarding_Status_{safe_name}.xlsx"

            st.download_button(
                label="⬇️ Download onboarding status (Excel)",
                data=excel_buffer,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        with mid:
            next_email = compute_next_email(df_check)
            if next_email:
                email_id, why = next_email
                st.info(f"Suggested next email: **{email_id}** (to progress: *{why}*)")
            else:
                st.success("No pending email-triggered items 🎉")

        with right:
            applicable = df_check[df_check["na"] == 0]
            total = len(applicable)
            done_count = int(applicable["done"].sum())
            st.progress(done_count / total if total else 0)
            st.write(f"Progress: **{done_count}/{total}** complete")

            model = derive_sales_model(df_check)
            st.write(f"{model['icon']} Sales model: {model['model']}")
            st.caption(model["hint"])

            # ✅ ADD THIS BLOCK RIGHT HERE
            st.divider()
            st.caption("Quick status")

            quick_map = [
                ("Signed service agreement", ("item", "Service agreement signed")),
                ("Signed customs representation agreement", ("item", "Customs representation agreement signed")),
                ("Signed fiscal representation agreement", ("item", "Fiscal representation agreement signed")),
                ("EORI number", ("item", "EORI confirmed / applied")),
                ("Insurance", ("item", "Insurance level confirmed")),
                ("WLI importer details", ("item", "Back-label importer chosen")),
                ("WLI portal", ("group", "D")),
                ("Internal handover", ("group", "E")),
            ]

            for label, rule in quick_map:
                kind, val = rule

                if kind == "item":
                    state = item_state(df_check, val)  # <- NEW 3-state logic
                    icon = {
                        "done": "✅",  # actually completed
                        "na": "➖",  # intentionally N/A (not required)
                        "open": "⬜"  # still to do
                    }[state]

                else:  # group (D, E, etc.)
                    ok = is_group_complete(df_check, val)
                    icon = "✅" if ok else "⬜"

                st.write(f"{icon} {label}")

        st.divider()

        # Update status with rules
        can_op, missing = can_set_operational(df_check)

        colA, colB = st.columns([1, 2])
        with colA:
            current_status = row["status"]
            new_status = st.selectbox("Status", STATUS_CHOICES, index=STATUS_CHOICES.index(current_status))
            if new_status == "Operational" and not can_op:
                st.warning("Operational requires these items first:\n- " + "\n- ".join(missing))

            if st.button("Save status"):
                if new_status == "Operational" and not can_op:
                    st.error("Cannot set Operational yet. Complete required items first.")
                else:
                    c.execute("UPDATE companies SET status = ? WHERE id = ?", (new_status, company_id))
                    c.commit()
                    st.success("Status updated ✅")

        with colB:
            company_notes = st.text_area("Company notes", value=row.get("notes", ""), height=100)
            if st.button("Save notes"):
                c.execute("UPDATE companies SET notes = ? WHERE id = ?", (company_notes, company_id))
                c.commit()
                st.success("Notes saved ✅")


        st.subheader("Checklist")

        # --- Filters ---
        fA, fB, fC, fD = st.columns([2, 1, 1, 1])
        with fA:
            item_search = st.text_input("Search checklist items", key=f"chk_search_{company_id}")
        with fB:
            show_open_only = st.checkbox("Open only", value=False, key=f"chk_open_{company_id}")
        with fC:
            show_email_only = st.checkbox("Email-linked only", value=False, key=f"chk_email_{company_id}")
        with fD:
            show_na = st.checkbox("Show N/A", value=True, key=f"chk_na_{company_id}")

        # --- Grouped checklist in expanders ---
        for grp in ["A", "B", "C", "D", "E"]:
            df_all_grp = df_check[df_check["grp"] == grp].copy()
            if df_all_grp.empty:
                continue

            # Group progress (applicable only)
            applicable_grp = df_all_grp[df_all_grp["na"] == 0]
            done_grp = applicable_grp[applicable_grp["done"] == 1]
            grp_done = int(len(done_grp))
            grp_total = int(len(applicable_grp))

            exp_key = f"exp_open_{company_id}_{grp}"
            is_open = st.session_state.get(exp_key, False)

            # 👇 REPLACE your old expander line with this
            with st.expander(GROUP_LABELS[grp], expanded=is_open):
                st.caption(f"{grp_done}/{grp_total} done")

                df_g = df_all_grp.copy()

                # Apply filters
                if item_search.strip():
                    s = item_search.strip().lower()
                    df_g = df_g[df_g["item"].str.lower().str.contains(s)]

                if show_email_only:
                    df_g = df_g[df_g["email_ref"].notna()]

                if show_open_only:
                    df_g = df_g[df_g["done"] == 0]

                if not show_na:
                    df_g = df_g[df_g["na"] == 0]

                if df_g.empty:
                    st.caption("No items match the current filters.")
                    continue

                # Render each item in a compact row
                for _, r in df_g.iterrows():
                    col_item, col_done, col_na, col_when, col_more = st.columns([6, 1, 1, 2, 2])

                    with col_item:
                        st.markdown(f"**{r['item']}**")
                        refs = []
                        if r["brochure_ref"]:
                            refs.append(f"Brochure {r['brochure_ref']}")
                        if r["email_ref"]:
                            refs.append(f"Email {r['email_ref']}")
                        if refs:
                            st.caption(" • ".join(refs))

                    with col_done:
                        done_val = st.checkbox(
                            "Done",
                            value=bool(r["done"]),
                            key=f"done_{company_id}_{r['item']}",
                            on_change=keep_group_open,
                            args=(company_id, grp),
                        )

                    with col_na:
                        na_val = st.checkbox(
                            "N/A",
                            value=bool(r.get("na", 0)),
                            key=f"na_{company_id}_{r['item']}",
                            on_change=keep_group_open,
                            args=(company_id, grp),
                        )

                    # Mutual exclusion
                    if na_val and done_val:
                        done_val = False
                        st.warning("Done + N/A can’t both be selected — keeping N/A.")

                    with col_when:
                        st.caption("Done at")

                        # Only allow editing if Done and not N/A
                        if done_val and not na_val:
                            existing_dt = parse_iso_z(r["done_at"]) if r["done_at"] else datetime.utcnow()
                            default_date = existing_dt.date()
                            default_time = existing_dt.time().replace(second=0, microsecond=0)

                            d = st.date_input(
                                "Done date",
                                value=default_date,
                                key=f"done_date_{company_id}_{r['item']}",
                                label_visibility="collapsed",
                            )
                            t = st.time_input(
                                "Done time",
                                value=default_time,
                                key=f"done_time_{company_id}_{r['item']}",
                                label_visibility="collapsed",
                            )

                            edited_done_at = iso_from_date_time(d, t)
                            st.caption(fmt_date(edited_done_at))
                        else:
                            edited_done_at = r["done_at"]
                            st.write(fmt_date(r["done_at"]))

                    # Notes + details tucked away
                    with col_more:
                        with st.popover("Details"):
                            st.write(f"**Brochure:** {r['brochure_ref'] or '—'}")
                            st.write(f"**Email:** {r['email_ref'] or '—'}")
                            note_val = st.text_area(
                                "Note",
                                value=r["note"] or "",
                                key=f"note_{company_id}_{r['item']}",
                                height=90,
                            )

                    # Determine changes
                    old_done = bool(r["done"])
                    old_na = bool(r.get("na", 0))
                    old_note = (r["note"] or "")

                    changed = (
                            (done_val != old_done)
                            or (na_val != old_na)
                            or (note_val != old_note)
                            or ((done_val and not na_val) and (edited_done_at != r["done_at"]))
                    )

                    if changed:
                        # Normalize states + done_at
                        if na_val:
                            done_val = False
                            done_at = None
                        else:
                            if done_val:
                                # If user set date/time, save that; otherwise fall back to now
                                done_at = edited_done_at or now_z()
                            else:
                                done_at = None

                        c.execute(
                            """
                            UPDATE checklist
                            SET done    = ?,
                                na      = ?,
                                done_at = ?,
                                note    = ?
                            WHERE company_id = ?
                              AND item = ?
                            """,
                            (
                                1 if done_val else 0,
                                1 if na_val else 0,
                                done_at,
                                note_val,
                                company_id,
                                r["item"],
                            ),
                        )
                        c.commit()

                        # Refresh checklist (so progress + gating uses latest)
                        df_check = load_checklist(c, company_id)

                        # Auto-status update (N/A-aware derive_status recommended)
                        current_status = pd.read_sql_query(
                            "SELECT status FROM companies WHERE id = ?",
                            c, params=(company_id,)
                        ).iloc[0]["status"]

                        auto_status = derive_status(current_status, df_check)
                        if auto_status != current_status:
                            c.execute("UPDATE companies SET status = ? WHERE id = ?", (auto_status, company_id))
                            c.commit()
                            st.toast(f"Status → {auto_status}", icon="ℹ️")
                        else:
                            st.toast("Saved", icon="✅")

                        # Refresh company row for header UI
                        df_company = load_company(c, company_id)
                        row = df_company.iloc[0]

        st.divider()

        st.subheader("Log an event (shows up in Timeline)")

        with st.form(f"add_event_{company_id}"):
            e1, e2 = st.columns([1, 1])
            with e1:
                event_date = st.date_input("Event date", key=f"ev_date_{company_id}")
            with e2:
                event_time = st.time_input("Event time", key=f"ev_time_{company_id}")

            event_type = st.selectbox(
                "Type",
                ["Meeting", "Warehouse visit", "Call", "Email", "Task", "Note"],
                key=f"ev_type_{company_id}"
            )

            title = st.text_input(
                "Title",
                placeholder="e.g. Call with customer / Send agreement / Follow-up email",
                key=f"ev_title_{company_id}"
            )

            detail = st.text_area(
                "Details",
                placeholder="What exactly needs to be done? Context, next steps, etc.",
                height=90,
                key=f"ev_detail_{company_id}"
            )

            # 🔔 Reminder options (ONLY for Tasks)
            send_reminder = False
            remind_minutes = 60

            if event_type == "Task":
                st.markdown("**🔔 Email reminder**")
                send_reminder = st.checkbox(
                    "Send email reminder before due time",
                    value=True,
                    key=f"rem_enable_{company_id}"
                )

                remind_minutes = st.number_input(
                    "Remind me (minutes before)",
                    min_value=5,
                    max_value=1440,
                    value=60,
                    step=5,
                    key=f"rem_minutes_{company_id}"
                )

            save_event = st.form_submit_button("Add event")

        if save_event:
            dt = datetime(
                event_date.year, event_date.month, event_date.day,
                event_time.hour, event_time.minute, 0
            )
            event_at_iso = dt.strftime("%Y-%m-%dT%H:%M:%SZ")

            # ⭐ THIS is your line, correctly integrated
            rm = int(remind_minutes) if (event_type == "Task" and send_reminder) else None

            add_event(
                c,
                company_id,
                event_at_iso,
                event_type,
                title.strip(),
                detail.strip(),
                remind_minutes=rm
            )

            st.success("Event added ✅ (will appear in Timeline & Tasks)")
            st.rerun()

        st.divider()

        st.subheader("Quick actions")

        a1, a2 = st.columns([1, 1])
        with a1:
            if st.button("Mark: Orders team informed"):
                c.execute("""
                    UPDATE checklist SET done = 1, done_at = ?, note = ?
                    WHERE company_id = ? AND item = ?
                """, (now_z(), "orders@wine-logistics.com notified", company_id, "Orders team informed"))
                c.commit()
                st.success("Orders team informed ✅")

                df_check = load_checklist(c, company_id)

                current_status = pd.read_sql_query(
                    "SELECT status FROM companies WHERE id = ?",
                    c, params=(company_id,)
                ).iloc[0]["status"]

                auto_status = derive_status(current_status, df_check)

                if auto_status != current_status:
                    c.execute(
                        "UPDATE companies SET status = ? WHERE id = ?",
                        (auto_status, company_id)
                    )
                    c.commit()
                    st.info(f"Status auto-updated to: {auto_status}")

            # ✅ REFRESH COMPANY DATA SO UI UPDATES
            df_company = load_company(c, company_id)
            row = df_company.iloc[0]

        with a2:
            if st.button("Mark: Warehouse informed"):
                c.execute("""
                    UPDATE checklist SET done = 1, done_at = ?, note = ?
                    WHERE company_id = ? AND item = ?
                """, (now_z(), "Warehouse (Hendrickx) notified", company_id, "Warehouse informed"))
                c.commit()
                st.success("Warehouse informed ✅")

                df_check = load_checklist(c, company_id)

                current_status = pd.read_sql_query(
                    "SELECT status FROM companies WHERE id = ?",
                    c, params=(company_id,)
                ).iloc[0]["status"]

                auto_status = derive_status(current_status, df_check)

                if auto_status != current_status:
                    c.execute(
                        "UPDATE companies SET status = ? WHERE id = ?",
                        (auto_status, company_id)
                    )
                    c.commit()
                    st.info(f"Status auto-updated to: {auto_status}")

            # ✅ REFRESH COMPANY DATA SO UI UPDATES
            df_company = load_company(c, company_id)
            row = df_company.iloc[0]

        st.caption("Next follow-up")

        existing = row.get("next_followup_at")
        existing_dt = parse_iso_z(existing) if existing else None

        dflt_date = (existing_dt.date() if existing_dt else datetime.utcnow().date())
        dflt_time = (
            existing_dt.time().replace(second=0, microsecond=0) if existing_dt else datetime.utcnow().time().replace(
                second=0, microsecond=0))

        fu1, fu2 = st.columns([1, 1])
        with fu1:
            fu_date = st.date_input("Follow-up date", value=dflt_date, key=f"fu_date_{company_id}")
        with fu2:
            fu_time = st.time_input("Follow-up time", value=dflt_time, key=f"fu_time_{company_id}")

        if st.button("Save follow-up", key=f"save_fu_{company_id}"):
            fu_iso = iso_from_date_time(fu_date, fu_time)
            c.execute("UPDATE companies SET next_followup_at = ? WHERE id = ?", (fu_iso, company_id))
            c.commit()
            st.toast("Follow-up saved", icon="🕘")
            st.rerun()

# ------------------- TAB 3 -------------------
if page == "📧 Email library":
    st.subheader("Email library (copy/paste)")
    st.caption("These are templates you can paste into Gmail. Personalize {name}, and where relevant {username}/{password}.")

    email_id = st.selectbox("Choose an email template", list(EMAIL_TEMPLATES.keys()))
    tpl = EMAIL_TEMPLATES[email_id]

    st.markdown(f"**Reference ID:** {email_id}")
    st.markdown(f"**Subject:** {tpl['subject']}")
    st.text_area("Body", value=tpl["body"], height=350)

    st.divider()
    st.subheader("Which checklist items does this email usually move?")
    # show checklist items tied to selected email
    tied = [x for x in CHECKLIST_ITEMS if x[3] == email_id]
    if not tied:
        st.write("No checklist items are linked to this email.")
    else:
        df = pd.DataFrame(tied, columns=["Group", "Checklist item", "Brochure ref", "Email ref"])
        df["Group"] = df["Group"].map(GROUP_LABELS)
        st.dataframe(df[["Group", "Checklist item", "Brochure ref"]], use_container_width=True, hide_index=True)

if page == "🕘 Timeline":
    st.subheader("Customer Timeline")

    df_companies = load_companies(c)
    if df_companies.empty:
        st.info("No companies yet.")
    else:
        # --- Select company ---
        options = df_companies[["id", "company_name"]].values.tolist()
        selected = st.selectbox(
            "Select a company",
            options,
            format_func=lambda x: f"{x[1]} (ID {x[0]})",
            key="timeline_company_select"
        )
        company_id = int(selected[0])

        df_company = load_company(c, company_id)
        df_check = load_checklist(c, company_id)
        row = df_company.iloc[0]

        # --- Header ---
        # --- Header ---
        st.write(f"**{row['company_name']}** — status: **{row['status']}**")

        col_left, col_right = st.columns([2, 1])

        # LEFT = Contacts (existing info)
        with col_left:
            df_contacts = load_contacts(c, company_id)

            if df_contacts.empty:
                # Legacy fallback
                st.write(f"**Contact:** {row['contact_name']} — {row['contact_email']}")
            else:
                st.write("**Contacts:**")
                for _, r in df_contacts.iterrows():
                    tag = " (primary)" if int(r["is_primary"]) == 1 else ""
                    role = f" — {r['role']}" if str(r.get("role", "")).strip() else ""
                    st.write(f"- {r['name']} — {r['email']}{role}{tag}")

            st.write(f"**Country:** {row['country']}")

        # RIGHT = Quick status snapshot
        with col_right:
            st.write("**Quick status**")

            quick_map = [
                ("Service agreement", ("item", "Service agreement signed")),
                ("Customs rep. agreement", ("item", "Customs representation agreement signed")),
                ("Fiscal rep. agreement", ("item", "Fiscal representation agreement signed")),
                ("EORI", ("item", "EORI confirmed / applied")),
                ("Insurance", ("item", "Insurance level confirmed")),
                ("Importer details", ("item", "Back-label importer chosen")),
                ("WLI portal", ("group", "D")),
                ("Internal handover", ("group", "E")),
            ]

            for label, rule in quick_map:
                kind, val = rule

                if kind == "item":
                    state = item_state(df_check, val)  # <- NEW 3-state logic
                    icon = {
                        "done": "✅",  # actually completed
                        "na": "➖",  # intentionally N/A (not required)
                        "open": "⬜"  # still to do
                    }[state]

                else:  # group (D, E, etc.)
                    ok = is_group_complete(df_check, val)
                    icon = "✅" if ok else "⬜"

                st.write(f"{icon} {label}")

        st.divider()

        # --- Build + render combined timeline ---
        df_custom_events = load_events(c, company_id)
        df_events = build_timeline(df_company, df_check, df_custom_events)

        timeline_df = df_events[["When", "Type", "Detail", "Email ref", "Note"]].copy()

        # ✅ 1) Render the timeline on the page (same HTML builder)
        render_vertical_timeline(timeline_df)

        st.divider()

        # ✅ 2) Build PDF ONLY when user clicks (avoids Playwright on every rerun)
        safe_name = "".join(ch for ch in row["company_name"] if ch.isalnum() or ch in (" ", "-", "_")).strip()
        pdf_name = f"WLI_Timeline_{safe_name}.pdf"

        if st.button("📄 Prepare PDF (same as timeline)", key=f"make_tl_pdf_{company_id}"):
            timeline_html = build_vertical_timeline_html(timeline_df)
            pdf_buffer = html_to_pdf_bytes(timeline_html)

            st.download_button(
                "⬇️ Download timeline (PDF)",
                data=pdf_buffer,
                file_name=pdf_name,
                mime="application/pdf",
                key=f"dl_tl_pdf_{company_id}",
            )

        st.divider()

        # --- Table view ---
        st.subheader("Timeline table")
        st.dataframe(
            df_events[["When", "Type", "Detail", "Email ref", "Note"]],
            use_container_width=True,
            hide_index=True
        )

        st.divider()

        st.subheader("Edit/delete manual events")

        df_custom_events = load_events(c, company_id)

        if df_custom_events.empty:
            st.caption("No manual events yet.")
        else:
            for _, ev in df_custom_events.sort_values("event_at", ascending=True).iterrows():
                ev_id = int(ev["id"])

                # Parse stored ISO to date/time defaults (fallback safe)
                dt = parse_iso_z(ev["event_at"]) or datetime.utcnow()
                ev_date_default = dt.date()
                ev_time_default = dt.time().replace(second=0, microsecond=0)

                with st.expander(f"{fmt_date(ev['event_at'])} — {ev['event_type']} — {ev['title'] or ''}".strip()):

                    # ✅ ADD THIS BLOCK HERE (top of expander)
                    if ev["event_type"] == "Task":
                        cb_key = f"edit_task_done_{company_id}_{ev_id}"
                        st.checkbox(
                            "Done",
                            value=int(ev.get("is_done", 0)) == 1,
                            key=cb_key,
                            on_change=toggle_task_done,
                            args=(ev_id, cb_key),
                        )

                    c1, c2 = st.columns([1, 1])

                    with c1:
                        new_date = st.date_input("Date", value=ev_date_default,
                                                 key=f"edit_ev_date_{company_id}_{ev_id}")
                    with c2:
                        new_time = st.time_input("Time", value=ev_time_default,
                                                 key=f"edit_ev_time_{company_id}_{ev_id}")

                    new_type = st.selectbox(
                        "Type",
                        ["Meeting", "Warehouse visit", "Call", "Email", "Task", "Note"],
                        index=["Meeting", "Warehouse visit", "Call", "Email", "Task", "Note"].index(ev["event_type"]) if
                        ev["event_type"] in ["Meeting", "Warehouse visit", "Call", "Email", "Task", "Note"] else 5,
                        key=f"edit_ev_type_{company_id}_{ev_id}",
                    )

                    new_title = st.text_input("Title", value=ev["title"] or "",
                                              key=f"edit_ev_title_{company_id}_{ev_id}")
                    new_detail = st.text_area("Details", value=ev["detail"] or "", height=90,
                                              key=f"edit_ev_detail_{company_id}_{ev_id}")

                    b1, b2 = st.columns([1, 1])
                    with b1:
                        if st.button("💾 Save changes", key=f"save_ev_{company_id}_{ev_id}"):
                            if not new_title.strip():
                                st.error("Title cannot be empty.")
                            else:
                                new_dt = datetime(new_date.year, new_date.month, new_date.day, new_time.hour,
                                                  new_time.minute, 0)
                                new_iso = new_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
                                update_event(c, ev_id, new_iso, new_type, new_title.strip(), new_detail.strip())
                                st.success("Saved ✅")
                                st.rerun()

                    with b2:
                        if st.button("🗑️ Delete", key=f"del_ev_{company_id}_{ev_id}"):
                            delete_event(c, ev_id)
                            st.success("Deleted ✅")
                            st.rerun()

        st.divider()

        # --- Human-readable activity feed ---
        st.subheader("Activity feed")
        for _, e in df_events.iterrows():
            when = fmt_date(e["When"])
            typ = e["Type"]
            detail = e["Detail"]
            email_ref = e.get("Email ref", "")
            note = (e.get("Note", "") or "").strip()

            suffix = f" ({email_ref})" if email_ref else ""
            st.write(f"- **{when}** — {typ}: {detail}{suffix}")
            if note:
                st.caption(note)

# ------------------- CALENDAR / ANALYTICS PAGE -------------------
# ------------------- CALENDAR / ANALYTICS PAGE -------------------
if page == "📅 Calendar / Analytics":
    st.subheader("New potential customers – Calendar view")

    # Load all companies
    df_comp, _ = load_dashboard_data(c)

    if df_comp.empty:
        st.info("No companies yet.")
    else:
        # Parse created_at to date
        df_comp["created_date"] = pd.to_datetime(
            df_comp["created_at"], errors="coerce"
        ).dt.date

        # Month selector (default = current month)
        today = datetime.utcnow().date()
        year = st.number_input("Year", min_value=2020, max_value=2100, value=today.year)
        month = st.selectbox(
            "Month",
            list(range(1, 13)),
            index=today.month - 1,
            format_func=lambda m: calendar.month_name[m]
        )

        # Filter to selected month
        start_date = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = date(year, month, last_day)

        df_month = df_comp[
            (df_comp["created_date"] >= start_date) &
            (df_comp["created_date"] <= end_date)
        ].copy()

        # KPI metrics
        k1, k2, k3 = st.columns(3)
        k1.metric("New prospects this month", len(df_month))

        if not df_month.empty:
            daily_counts = df_month.groupby("created_date").size()
            k2.metric("Busiest day", int(daily_counts.max()))
            k3.metric("Active days", int((daily_counts > 0).sum()))
        else:
            k2.metric("Busiest day", 0)
            k3.metric("Active days", 0)

        st.divider()

        st.write(f"### {calendar.month_name[month]} {year}")

        # Build calendar grid
        cal = calendar.Calendar(firstweekday=0)  # Monday
        weeks = cal.monthdatescalendar(year, month)

        # Count companies per day
        counts = df_month.groupby("created_date").size().to_dict()

        # Weekday header
        headers = st.columns(7)
        for i, d in enumerate(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]):
            headers[i].markdown(f"**{d}**")

        # Render weeks
        for week in weeks:
            cols = st.columns(7)
            for i, day in enumerate(week):
                in_month = day.month == month
                count = counts.get(day, 0)

                if not in_month:
                    cols[i].markdown("<div style='opacity:0.3'> </div>", unsafe_allow_html=True)
                    continue

                # Color intensity based on number of prospects
                if count == 0:
                    bg = "#f5f5f5"
                elif count == 1:
                    bg = "#c8e6c9"
                elif count == 2:
                    bg = "#81c784"
                else:
                    bg = "#2e7d32"

                cols[i].markdown(
                    f"""
                    <div style="
                        border:1px solid rgba(0,0,0,0.1);
                        border-radius:12px;
                        padding:8px;
                        min-height:70px;
                        background:{bg};
                    ">
                        <div style="font-size:12px; opacity:0.7;">{day.day}</div>
                        <div style="font-size:22px; font-weight:700;">{count}</div>
                        <div style="font-size:11px;">prospects</div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

if page == "📇 Companies":
    st.subheader("Companies")

    df_companies = pd.read_sql_query("""
        SELECT id, created_at, company_name, country, contact_name, contact_email, status, next_followup_at
        FROM companies
        ORDER BY id DESC
    """, c)

    if df_companies.empty:
        st.info("No companies yet.")
    else:
        q = st.text_input("Search (company / email / country / status)")
        dff = df_companies.copy()
        if q.strip():
            s = q.strip().lower()
            dff = dff[
                dff["company_name"].str.lower().str.contains(s, na=False)
                | dff["contact_email"].str.lower().str.contains(s, na=False)
                | dff["country"].str.lower().str.contains(s, na=False)
                | dff["status"].str.lower().str.contains(s, na=False)
            ]

        # Show a compact table
        dff_show = dff.copy()
        dff_show["Created"] = dff_show["created_at"].apply(lambda x: fmt_date(x, with_time=False))
        dff_show["Next follow-up"] = dff_show["next_followup_at"].apply(lambda x: fmt_date(x) if x else "—")

        st.dataframe(
            dff_show[["id", "company_name", "status", "country", "Created", "Next follow-up"]],
            use_container_width=True,
            hide_index=True
        )

        # Select + open
        options = dff[["id", "company_name"]].values.tolist()
        selected = st.selectbox(
            "Open company record",
            options,
            format_func=lambda x: f"{x[1]} (ID {x[0]})"
        )

        if st.button("🏢 Open", use_container_width=True):
            set_active_company(int(selected[0]), go_page="🏢 Company record")
            st.rerun()

if page == "🏢 Company record":
    st.subheader("Company record")

    df_companies = load_companies(c)
    if df_companies.empty:
        st.info("No companies yet.")
        st.stop()

    # Pick company: default to active_company_id if set
    options = df_companies[["id", "company_name"]].values.tolist()

    default_index = 0
    if st.session_state.active_company_id is not None:
        for i, (cid, _) in enumerate(options):
            if int(cid) == int(st.session_state.active_company_id):
                default_index = i
                break

    selected = st.selectbox(
        "Select a company",
        options,
        index=default_index,
        format_func=lambda x: f"{x[1]} (ID {x[0]})",
        key="company_record_select"
    )
    company_id = int(selected[0])
    st.session_state.active_company_id = company_id  # keep synced

    df_company = load_company(c, company_id)
    df_check = load_checklist(c, company_id)
    model = derive_sales_model(df_check)
    row = df_company.iloc[0]

    # --- Header snapshot ---
    left, mid, right = st.columns([2.5, 2, 2])

    with left:
        st.markdown(f"## {row['company_name']}")
        st.write(f"**Status:** {row['status']}")
        st.write(f"**Country:** {row['country']}")
        st.write(f"**Created:** {fmt_date(row['created_at'], with_time=False)}")

    with mid:
        # Progress
        prog = calc_progress(df_check)
        st.metric("Checklist progress", f"{prog}%")
        st.progress(prog / 100)

        # Sales model (derived from customs/fiscal logic)
        model = derive_sales_model(df_check)
        st.metric("Sales model", f"{model['icon']} {model['model']}")
        st.caption(model["hint"])

        st.divider()

        # Suggested next email
        nxt = compute_next_email(df_check)
        if nxt:
            email_id, why = nxt
            st.info(f"Suggested next email: **{email_id}** (for *{why}*)")
        else:
            st.success("No pending email-triggered items 🎉")

    with right:
        st.write("**Quick status**")
        quick_map = [
            ("Service agreement", ("item", "Service agreement signed")),
            ("Customs rep. agreement", ("item", "Customs representation agreement signed")),
            ("Fiscal rep. agreement", ("item", "Fiscal representation agreement signed")),
            ("EORI", ("item", "EORI confirmed / applied")),
            ("Insurance", ("item", "Insurance level confirmed")),
            ("Importer details", ("item", "Back-label importer chosen")),
            ("WLI portal", ("group", "D")),
            ("Internal handover", ("group", "E")),
        ]
        for label, rule in quick_map:
            kind, val = rule
            ok = is_item_satisfied(df_check, val) if kind == "item" else is_group_complete(df_check, val)
            st.write(("✅ " if ok else "⬜ ") + label)

    st.divider()

    # --- Actions row ---
    a1, a2, a3 = st.columns([1.3, 1.3, 2])
    with a1:
        excel_buffer = generate_onboarding_excel(row, df_check)
        safe_name = "".join(ch for ch in row["company_name"] if ch.isalnum() or ch in (" ", "-", "_")).strip()
        file_name = f"WLI_Onboarding_Status_{safe_name}.xlsx"
        st.download_button(
            label="⬇️ Download status (Excel)",
            data=excel_buffer,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with a2:
        if st.button("📋 Go to checklist", use_container_width=True):
            set_active_company(company_id, go_page="📋 Pipeline / Checklist")

    with a3:
        st.caption("Tip: use this page for editing + deletion. Use Pipeline for checklist work.")

    st.divider()

    # --- Contacts + follow-up + notes ---
    colA, colB = st.columns([2, 2])

    with colA:
        st.subheader("Contacts")
        df_contacts = load_contacts(c, company_id)

        if df_contacts.empty:
            st.caption("No contacts found (legacy record).")
        else:
            st.dataframe(df_contacts, use_container_width=True, hide_index=True)

        with st.expander("➕ Add contact"):
            with st.form(f"add_contact_form_{company_id}"):
                n = st.text_input("Name")
                e = st.text_input("Email")
                r = st.text_input("Role")
                primary = st.checkbox("Set as primary")
                addc = st.form_submit_button("Add contact")

            if addc:
                if not n.strip() or not e.strip():
                    st.error("Name + Email required.")
                else:
                    if primary:
                        # unset other primaries
                        c.execute("UPDATE company_contacts SET is_primary = 0 WHERE company_id = ?", (company_id,))
                        c.commit()
                    add_contact(c, company_id, n, e, r, is_primary=1 if primary else 0)
                    st.success("Contact added ✅")
                    st.rerun()

    with colB:
        st.subheader("Next follow-up")

        existing = row.get("next_followup_at")
        existing_dt = parse_iso_z(existing) if existing else None

        dflt_date = existing_dt.date() if existing_dt else datetime.utcnow().date()
        dflt_time = (existing_dt.time().replace(second=0, microsecond=0)
                     if existing_dt else datetime.utcnow().time().replace(second=0, microsecond=0))

        fu1, fu2 = st.columns(2)
        with fu1:
            fu_date = st.date_input("Follow-up date", value=dflt_date, key=f"cr_fu_date_{company_id}")
        with fu2:
            fu_time = st.time_input("Follow-up time", value=dflt_time, key=f"cr_fu_time_{company_id}")

        if st.button("Save follow-up", key=f"cr_save_fu_{company_id}", use_container_width=True):
            fu_iso = iso_from_date_time(fu_date, fu_time)
            c.execute("UPDATE companies SET next_followup_at = ? WHERE id = ?", (fu_iso, company_id))
            c.commit()
            st.toast("Follow-up saved", icon="🕘")
            st.rerun()

        st.divider()
        st.subheader("Notes")

        notes_val = st.text_area("Company notes", value=row.get("notes", "") or "", height=140, key=f"cr_notes_{company_id}")
        if st.button("Save notes", key=f"cr_save_notes_{company_id}", use_container_width=True):
            c.execute("UPDATE companies SET notes = ? WHERE id = ?", (notes_val.strip(), company_id))
            c.commit()
            st.success("Notes saved ✅")
            st.rerun()

    st.divider()

    # --- Edit core company fields ---
    st.subheader("Company details")

    with st.expander("✏️ Edit company details", expanded=False):
        with st.form(f"cr_edit_company_{company_id}"):
            c1, c2 = st.columns(2)
            with c1:
                new_company_name = st.text_input("Company name", value=row["company_name"] or "")
                new_country = st.text_input("Country", value=row["country"] or "")
            with c2:
                new_status = st.selectbox("Status", STATUS_CHOICES, index=STATUS_CHOICES.index(row["status"]))
                new_contact_name = st.text_input("Legacy contact name", value=row["contact_name"] or "")
                new_contact_email = st.text_input("Legacy contact email", value=row["contact_email"] or "")
            save = st.form_submit_button("Save")

        if save:
            if not new_company_name.strip():
                st.error("Company name cannot be empty.")
            else:
                c.execute("""
                    UPDATE companies
                    SET company_name = ?, country = ?, status = ?, contact_name = ?, contact_email = ?
                    WHERE id = ?
                """, (new_company_name.strip(), new_country.strip(), new_status, new_contact_name.strip(), new_contact_email.strip(), company_id))
                c.commit()

                # Keep primary contact synced if you want (optional, but you already do this elsewhere)
                c.execute("""
                    UPDATE company_contacts
                    SET name = ?, email = ?
                    WHERE company_id = ? AND is_primary = 1
                """, (new_contact_name.strip(), new_contact_email.strip(), company_id))
                c.commit()

                st.success("Updated ✅")
                st.rerun()

    st.divider()

    # --- Danger zone ---
    st.subheader("Danger zone")
    with st.expander("🗑️ Delete this company record", expanded=False):
        st.warning("This will permanently delete the company AND all related checklist items, events, and contacts.")
        confirm_text = st.text_input("Type DELETE to confirm", key=f"cr_confirm_delete_{company_id}")
        if st.button("🗑️ Delete company", key=f"cr_delete_btn_{company_id}", use_container_width=True):
            if confirm_text.strip() != "DELETE":
                st.error("Confirmation text did not match.")
            else:
                delete_company(c, company_id)
                st.success("Company deleted ✅")
                st.session_state.active_company_id = None
                st.session_state.nav_request = "📇 Companies"
                st.rerun()

